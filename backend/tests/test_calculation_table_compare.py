"""Declared LIRA xlsx/docx table compare — сверка only, not a solver."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from aerobim.core.config.settings import Settings
from aerobim.domain.calculation_evidence import CALCULATION_CORRECTNESS_CLAIM
from aerobim.domain.calculation_table_compare import (
    DeclaredCalcRow,
    compare_declared_tables,
    table_compare_honesty_snapshot,
    table_digest,
)
from aerobim.domain.ifc_streaming_design import (
    DEFAULT_ANALYZE_IFC_BYTES,
    streaming_design_snapshot,
)
from aerobim.domain.models import RequirementSource, SourceKind, ValidationRequest
from aerobim.domain.samolet_mvp_answers import samolet_mvp_answers_payload
from aerobim.infrastructure.adapters.spreadsheet_load_evidence_adapter import (
    SpreadsheetLoadEvidenceAdapter,
    compare_office_tables,
    extract_declared_rows,
)


def _repo() -> Path:
    return Path(__file__).resolve().parents[2]


class CalculationTableCompareTests(unittest.TestCase):
    def test_mismatch_and_digest_stable(self) -> None:
        calc = (
            DeclaredCalcRow("L1", "snow", "10", "kN"),
            DeclaredCalcRow("L2", "wind", "5", "kN"),
        )
        bim = (
            DeclaredCalcRow("L1", "snow", "10", "kN"),
            DeclaredCalcRow("L2", "wind", "8", "kN"),
        )
        result = compare_declared_tables(calc, bim)
        self.assertFalse(result.all_match)
        self.assertEqual(result.solver, "not_implemented")
        self.assertEqual(result.claim, CALCULATION_CORRECTNESS_CLAIM)
        self.assertFalse(result.closes_rt001)
        by_id = {item.field_id: item for item in result.fields}
        self.assertEqual(by_id["L1"].outcome, "MATCH")
        self.assertEqual(by_id["L2"].outcome, "MISMATCH")
        self.assertEqual(result.calc_digest, table_digest(calc))
        self.assertNotEqual(result.calc_digest, result.bim_digest)

    def test_all_match_still_not_solver(self) -> None:
        rows = (DeclaredCalcRow("A", "area", "120.5", "m2"),)
        result = compare_declared_tables(rows, rows)
        self.assertTrue(result.all_match)
        self.assertEqual(result.solver, "not_implemented")

    def test_duplicate_ids_are_not_match(self) -> None:
        calc = (
            DeclaredCalcRow("L1", "a", "1", "kN"),
            DeclaredCalcRow("L1", "b", "1", "kN"),
        )
        bim = (DeclaredCalcRow("L1", "a", "1", "kN"),)
        result = compare_declared_tables(calc, bim)
        self.assertFalse(result.all_match)
        self.assertIn("L1", result.duplicate_ids)

    def test_committed_json_fixture(self) -> None:
        path = _repo() / "samples" / "calculations" / "lira-declared-vs-rd.json"
        payload = json.loads(path.read_text(encoding="utf-8"))
        self.assertTrue(payload["not_solver"])
        self.assertFalse(payload["closes_rt001"])
        calc = tuple(DeclaredCalcRow(**row) for row in payload["calc_rows"])
        bim = tuple(DeclaredCalcRow(**row) for row in payload["bim_rows"])
        result = compare_declared_tables(calc, bim)
        self.assertFalse(result.all_match)

    def test_honesty_snapshot_and_capabilities(self) -> None:
        snap = table_compare_honesty_snapshot()
        self.assertEqual(snap["solver"], "not_implemented")
        self.assertEqual(snap["pdf_tables"], "fragile")
        payload = samolet_mvp_answers_payload()
        compare = payload["calculation_table_compare"]
        self.assertIsInstance(compare, dict)
        self.assertEqual(compare["native_lir"], "not_implemented")
        self.assertFalse(compare["closes_rt002"])


class OfficeDeclaredTableAdapterTests(unittest.TestCase):
    def test_xlsx_and_docx_roundtrip(self) -> None:
        try:
            from docx import Document
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("python-docx/openpyxl not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xlsx = root / "lira.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(["id", "label", "value", "unit"])
            ws.append(["L1", "snow", "10", "kN"])
            ws.append(["L2", "wind", "5", "kN"])
            wb.save(xlsx)
            docx_path = root / "rd.docx"
            doc = Document()
            table = doc.add_table(rows=3, cols=4)
            headers = ("id", "label", "value", "unit")
            values = (("L1", "snow", "10", "kN"), ("L2", "wind", "8", "kN"))
            for col, header in enumerate(headers):
                table.cell(0, col).text = header
            for row_idx, row in enumerate(values, start=1):
                for col, cell in enumerate(row):
                    table.cell(row_idx, col).text = cell
            doc.save(docx_path)
            calc = extract_declared_rows(xlsx)
            bim = extract_declared_rows(docx_path)
            self.assertEqual(calc.status, "ok")
            self.assertEqual(bim.status, "ok")
            result = compare_office_tables(xlsx, docx_path)
            self.assertFalse(result.all_match)
            self.assertEqual(result.solver, "not_implemented")

    def test_native_lir_and_pdf_fail_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            lir = Path(tmp) / "model.lir"
            lir.write_bytes(b"LIRA-BIN")
            pdf = Path(tmp) / "note.pdf"
            pdf.write_bytes(b"%PDF-1.4\n")
            self.assertEqual(extract_declared_rows(lir).status, "native_lir_not_implemented")
            self.assertEqual(extract_declared_rows(pdf).status, "pdf_fragile")
            adapter = SpreadsheetLoadEvidenceAdapter()
            ifc = Path(tmp) / "m.ifc"
            ifc.write_text("ISO-10303-21;", encoding="utf-8")
            lir_issues = adapter.verify(
                ValidationRequest(
                    request_id="lir",
                    ifc_path=ifc,
                    requirement_source=RequirementSource(text="R|IFCWALL|P|T|1"),
                    calculation_source=RequirementSource(
                        text="",
                        path=lir,
                        source_kind=SourceKind.CALCULATION,
                    ),
                )
            )
            self.assertTrue(any(i.rule_id == "AEROBIM-LIRA-NATIVE" for i in lir_issues))
            pdf_issues = adapter.verify(
                ValidationRequest(
                    request_id="pdf",
                    ifc_path=ifc,
                    requirement_source=RequirementSource(text="R|IFCWALL|P|T|1"),
                    calculation_source=RequirementSource(
                        text="",
                        path=pdf,
                        source_kind=SourceKind.CALCULATION,
                    ),
                )
            )
            self.assertTrue(any(i.rule_id == "AEROBIM-LIRA-PDF" for i in pdf_issues))


class IfcStreamingDesignTests(unittest.TestCase):
    def test_snapshot_does_not_raise_cap(self) -> None:
        snap = streaming_design_snapshot()
        self.assertEqual(snap["streaming_parser"], "designed_not_implemented")
        self.assertEqual(snap["disk_r_tree"], "designed_not_implemented")
        self.assertFalse(snap["raises_default_cap"])
        self.assertEqual(snap["spatial_index_json_sidecar"], "dump_only")
        self.assertEqual(snap["default_analyze_bytes"], DEFAULT_ANALYZE_IFC_BYTES)
        self.assertEqual(DEFAULT_ANALYZE_IFC_BYTES, 256 * 1024 * 1024)
        settings = Settings(
            application_name="aerobim-test",
            environment="test",
            host="127.0.0.1",
            port=8080,
            storage_dir=Path("."),
            debug=True,
        )
        self.assertEqual(settings.max_ifc_bytes, DEFAULT_ANALYZE_IFC_BYTES)
        payload = samolet_mvp_answers_payload()
        streaming = payload["ifc_streaming"]
        self.assertIsInstance(streaming, dict)
        self.assertFalse(streaming["raises_default_cap"])
        self.assertTrue(streaming["in_memory_spatial_index"])


if __name__ == "__main__":
    unittest.main()
